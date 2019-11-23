from bondMathLib import *
import openpyxl
from openpyxl.styles import Font 

if __name__ == "__main__":
    path = "C:\\Users\\Mike\\Desktop\\BondInfo.xlsx"

    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active
    
    years = sheet_obj.cell(row = 2, column = 1).value
    couponRate = sheet_obj.cell(row = 2, column = 2).value
    BondPrice = sheet_obj.cell(row = 2, column = 3).value
    compounding = sheet_obj.cell(row = 2, column = 4).value

    resultYield = computeYield(years,couponRate,BondPrice,compounding)
    print("the yield of a 3 year bond with coupon of 8% valued at 104 is " + str(resultYield))
    cell = sheet_obj.cell(row = 4, column = 1)
    cell.value = "Yield"
    cell.font = Font(size = 16, bold = True,underline = 'single', name = 'Times New Roman')
    sheet_obj.cell(row = 5, column = 1).value = resultYield

    resultDuration = computeDuration(years,couponRate,BondPrice,compounding)
    print("the Duration of a 3 year bond with coupon of 8% valued at 104 is " + str(resultDuration))
    cell = sheet_obj.cell(row = 4, column = 2)
    cell.value = "Duration"
    cell.font = Font(size = 16, bold = True,underline = 'single', name = 'Times New Roman')
    sheet_obj.cell(row = 5, column = 2).value = resultDuration

    resultDV01 = computeDV01(BondPrice,resultDuration)
    print("the DV01 for this bond is " + str(resultDV01))
    cell = sheet_obj.cell(row = 4, column = 3)
    cell.value = "DV01"
    cell.font = Font(size = 16, bold = True,underline = 'single', name = 'Times New Roman')
    sheet_obj.cell(row = 5, column = 3).value = resultDV01    

    wb_obj.save(path)

